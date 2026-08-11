"""Импорт и экспорт справочника сотрудников в безопасном XLSX-формате."""

from collections import Counter
from io import BytesIO
from zipfile import BadZipFile, ZipFile, is_zipfile

import psycopg2.extras
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from db import transaction

SHEET_NAME = 'Сотрудники'
LOOKUP_SHEET_NAME = 'Справочники'
HEADERS = ('Контрагент', 'Подразделение', 'Должность', 'Статус')
ALLOWED_STATUSES = ('Работает', 'Уволен')
UNASSIGNED_DEPARTMENT = 'Без подразделения'
MAX_FILE_SIZE = 10 * 1024 * 1024
MAX_UNCOMPRESSED_SIZE = 40 * 1024 * 1024
MAX_ROWS = 5000
MAX_COLUMNS = 20
MAX_TEXT_LENGTH = 500


class EmployeeWorkbookError(ValueError):
    """Понятная пользователю ошибка структуры или данных книги."""


def build_employee_summary(rows, department_order=()):
    """Считает распределение сотрудников по подразделениям и статусам."""
    counts = Counter()
    known_departments = []
    known_statuses = []

    for row in rows:
        department = _clean_text(row.get('department')) or UNASSIGNED_DEPARTMENT
        status = _clean_text(row.get('status')) or ALLOWED_STATUSES[0]
        counts[(department, status)] += 1
        known_departments.append(department)
        known_statuses.append(status)

    department_rank = {
        department: index
        for index, department in enumerate(_unique(department_order))
    }
    departments = sorted(
        set(known_departments),
        key=lambda department: (
            department == UNASSIGNED_DEPARTMENT,
            department_rank.get(department, len(department_rank)),
            department.casefold(),
        ),
    )
    extra_statuses = sorted(
        set(known_statuses) - set(ALLOWED_STATUSES),
        key=str.casefold,
    )
    statuses = list(ALLOWED_STATUSES) + extra_statuses
    totals = {
        status: sum(counts[(department, status)] for department in departments)
        for status in statuses
    }

    return {
        'statuses': statuses,
        'rows': [
            {
                'department': department,
                'counts': {
                    status: counts[(department, status)] for status in statuses
                },
                'total': sum(counts[(department, status)] for status in statuses),
            }
            for department in departments
        ],
        'totals': totals,
        'grand_total': sum(totals.values()),
    }


def filter_employee_rows(rows, department=None, status=None):
    """Фильтрует строки теми же значениями, которые показаны в сводной."""
    filtered = []
    for row in rows:
        row_department = _clean_text(row.get('department')) or UNASSIGNED_DEPARTMENT
        row_status = _clean_text(row.get('status')) or ALLOWED_STATUSES[0]
        if department is not None and row_department != department:
            continue
        if status is not None and row_status != status:
            continue
        filtered.append(row)
    return filtered


def _clean_text(value):
    if value is None:
        return ''
    if isinstance(value, bool):
        return 'TRUE' if value else 'FALSE'
    text = str(value).strip()
    if len(text) > MAX_TEXT_LENGTH:
        raise EmployeeWorkbookError(
            f'Значение длиннее {MAX_TEXT_LENGTH} символов: {text[:40]}…'
        )
    # При экспорте потенциальные формулы получают один защитный апостроф.
    if len(text) > 1 and text[0] == "'" and text[1] in '=+@-':
        text = text[1:]
    return text


def _excel_safe(value):
    text = _clean_text(value)
    return "'" + text if text.startswith(('=', '+', '@', '-')) else text


def _unique(values):
    seen = set()
    result = []
    for value in values:
        text = _clean_text(value)
        if text and text not in seen:
            seen.add(text)
            result.append(text)
    return result


def build_employee_workbook(rows, department_options=()):
    """Возвращает готовый XLSX со справочником и подсказками для пользователя."""
    rows = list(rows)
    workbook = Workbook()
    workbook.iso_dates = True
    sheet = workbook.active
    sheet.title = SHEET_NAME
    sheet.freeze_panes = 'A2'
    sheet.sheet_view.showGridLines = False

    header_fill = PatternFill('solid', fgColor='24344D')
    header_font = Font(color='FFFFFF', bold=True)
    key_fill = PatternFill('solid', fgColor='E9ECEF')
    editable_fill = PatternFill('solid', fgColor='EAF3FF')
    warning_fill = PatternFill('solid', fgColor='FFF3CD')
    thin_border = Border(bottom=Side(style='thin', color='D9DEE7'))

    for column, header in enumerate(HEADERS, 1):
        cell = sheet.cell(row=1, column=column, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    sheet.row_dimensions[1].height = 28

    for row_number, row in enumerate(rows, 2):
        values = (
            _excel_safe(row.get('contragent')),
            _excel_safe(row.get('department')),
            _excel_safe(row.get('position')),
            _excel_safe(row.get('status') or ALLOWED_STATUSES[0]),
        )
        for column, value in enumerate(values, 1):
            cell = sheet.cell(row=row_number, column=column, value=value)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = thin_border
            cell.fill = key_fill if column == 1 else editable_fill
            cell.protection = Protection(locked=column == 1)

    final_row = max(2, len(rows) + 1)
    if not rows:
        for column in range(1, len(HEADERS) + 1):
            sheet.cell(row=2, column=column, value='')

    sheet.column_dimensions['A'].width = 38
    sheet.column_dimensions['B'].width = 28
    sheet.column_dimensions['C'].width = 32
    sheet.column_dimensions['D'].width = 16
    sheet.auto_filter.ref = f'A1:D{final_row}'
    table = Table(displayName='EmployeesDirectory', ref=f'A1:D{final_row}')
    table.tableStyleInfo = TableStyleInfo(
        name='TableStyleMedium2', showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    sheet.add_table(table)
    sheet['A1'].comment = Comment(
        'Ключ сотрудника. Не изменяйте это поле: по нему строки сопоставляются при импорте.',
        'Alfa Top Report',
    )

    lookup = workbook.create_sheet(LOOKUP_SHEET_NAME)
    departments = _unique(
        list(department_options) + [row.get('department') for row in rows]
    )
    for row_number, department in enumerate(departments or [''], 1):
        lookup.cell(row=row_number, column=1, value=_excel_safe(department))
    for row_number, status in enumerate(ALLOWED_STATUSES, 1):
        lookup.cell(row=row_number, column=2, value=status)
    lookup.sheet_state = 'hidden'

    if departments:
        workbook.defined_names.add(DefinedName(
            'EmployeeDepartments',
            attr_text=f"'{LOOKUP_SHEET_NAME}'!$A$1:$A${len(departments)}",
        ))
        department_validation = DataValidation(
            type='list',
            formula1='EmployeeDepartments',
            allow_blank=True,
            errorStyle='warning',
            errorTitle='Новое подразделение',
            error='Можно оставить введённое значение или выбрать подразделение из списка.',
        )
        department_validation.showErrorMessage = True
        sheet.add_data_validation(department_validation)
        department_validation.add(f'B2:B{MAX_ROWS + 1}')

    workbook.defined_names.add(DefinedName(
        'EmployeeStatuses',
        attr_text=f"'{LOOKUP_SHEET_NAME}'!$B$1:$B${len(ALLOWED_STATUSES)}",
    ))
    status_validation = DataValidation(
        type='list',
        formula1='EmployeeStatuses',
        allow_blank=False,
        errorStyle='stop',
        errorTitle='Некорректный статус',
        error='Выберите «Работает» или «Уволен».',
    )
    status_validation.showErrorMessage = True
    sheet.add_data_validation(status_validation)
    status_validation.add(f'D2:D{MAX_ROWS + 1}')
    sheet.conditional_formatting.add(
        f'D2:D{MAX_ROWS + 1}',
        FormulaRule(formula=['D2="Уволен"'], fill=warning_fill),
    )

    instructions = workbook.create_sheet('Инструкция')
    instructions.sheet_view.showGridLines = False
    instructions.column_dimensions['A'].width = 105
    instructions['A1'] = 'Как обновить справочник сотрудников'
    instructions['A1'].font = Font(size=16, bold=True, color='24344D')
    instructions['A3'] = (
        '1. На листе «Сотрудники» изменяйте только голубые столбцы: '
        '«Подразделение», «Должность» и «Статус».'
    )
    instructions['A4'] = '2. Не меняйте «Контрагент»: это ключ для поиска сотрудника в системе.'
    instructions['A5'] = '3. Пустое подразделение или должность очистит соответствующее поле в системе.'
    instructions['A6'] = '4. Сохраните файл в формате XLSX и загрузите его на странице справочника.'
    instructions['A8'] = (
        'Импорт сначала проверяет весь файл и только затем применяет изменения. '
        'Сотрудники, которых нет в файле, остаются без изменений.'
    )
    for row_number in (3, 4, 5, 6, 8):
        instructions.cell(row=row_number, column=1).alignment = Alignment(wrap_text=True, vertical='top')
        instructions.row_dimensions[row_number].height = 34

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _validate_container(payload):
    if not payload:
        raise EmployeeWorkbookError('Файл пустой.')
    if len(payload) > MAX_FILE_SIZE:
        raise EmployeeWorkbookError('Файл больше 10 МБ.')
    stream = BytesIO(payload)
    if not is_zipfile(stream):
        raise EmployeeWorkbookError('Нужен настоящий файл Excel в формате XLSX.')
    stream.seek(0)
    try:
        with ZipFile(stream) as archive:
            total_size = sum(item.file_size for item in archive.infolist())
            if total_size > MAX_UNCOMPRESSED_SIZE:
                raise EmployeeWorkbookError('Содержимое файла слишком большое.')
    except BadZipFile as exc:
        raise EmployeeWorkbookError('Файл XLSX повреждён.') from exc


def parse_employee_workbook(payload):
    """Проверяет книгу целиком и возвращает нормализованные строки."""
    _validate_container(payload)
    try:
        workbook = load_workbook(BytesIO(payload), read_only=True, data_only=False, keep_links=False)
    except Exception as exc:
        raise EmployeeWorkbookError('Не удалось прочитать файл XLSX.') from exc

    try:
        if SHEET_NAME not in workbook.sheetnames:
            raise EmployeeWorkbookError(f'В файле нет листа «{SHEET_NAME}».')
        sheet = workbook[SHEET_NAME]
        if sheet.max_column > MAX_COLUMNS:
            raise EmployeeWorkbookError('В книге слишком много столбцов.')
        if sheet.max_row > MAX_ROWS + 1:
            raise EmployeeWorkbookError(f'В книге больше {MAX_ROWS} сотрудников.')

        header_cells = next(sheet.iter_rows(min_row=1, max_row=1), ())
        headers = [_clean_text(cell.value) for cell in header_cells]
        positions = {}
        for header in HEADERS:
            if headers.count(header) > 1:
                raise EmployeeWorkbookError(f'Столбец «{header}» встречается несколько раз.')
            try:
                positions[header] = headers.index(header)
            except ValueError as exc:
                raise EmployeeWorkbookError(f'Не найден обязательный столбец «{header}».') from exc

        errors = []
        employees = []
        seen = set()
        max_position = max(positions.values())
        for row_number, cells in enumerate(sheet.iter_rows(min_row=2), 2):
            values = list(cells)
            if len(values) <= max_position:
                values.extend([None] * (max_position + 1 - len(values)))
            selected = [values[positions[header]] for header in HEADERS]
            if all(cell is None or _clean_text(getattr(cell, 'value', None)) == '' for cell in selected):
                continue
            if any(getattr(cell, 'data_type', None) == 'f' for cell in selected if cell is not None):
                errors.append(f'строка {row_number}: формулы запрещены')
                continue
            try:
                contragent = _clean_text(selected[0].value)
                department = _clean_text(selected[1].value)
                position = _clean_text(selected[2].value)
                status = _clean_text(selected[3].value)
            except EmployeeWorkbookError as exc:
                errors.append(f'строка {row_number}: {exc}')
                continue
            if not contragent:
                errors.append(f'строка {row_number}: не заполнен «Контрагент»')
            elif contragent in seen:
                errors.append(f'строка {row_number}: сотрудник «{contragent}» встречается повторно')
            elif status not in ALLOWED_STATUSES:
                errors.append(
                    f'строка {row_number}: статус должен быть «Работает» или «Уволен»'
                )
            else:
                seen.add(contragent)
                employees.append({
                    'contragent': contragent,
                    'department': department or None,
                    'position': position or None,
                    'status': status,
                })
            if len(errors) >= 20:
                break

        if errors:
            suffix = ' Проверьте и остальные строки.' if len(errors) >= 20 else ''
            raise EmployeeWorkbookError('; '.join(errors) + suffix)
        if not employees:
            raise EmployeeWorkbookError('На листе «Сотрудники» нет данных для импорта.')
        return employees
    finally:
        workbook.close()


def apply_employee_updates(employees):
    """Атомарно обновляет только существующих сотрудников."""
    names = [employee['contragent'] for employee in employees]
    with transaction() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cursor:
            cursor.execute(
                '''SELECT contragent, department, position, status
                   FROM reporting.employees
                   WHERE contragent = ANY(%s)
                   FOR UPDATE''',
                (names,),
            )
            current = {row['contragent']: dict(row) for row in cursor.fetchall()}
            missing = [name for name in names if name not in current]
            if missing:
                preview = ', '.join(missing[:5])
                if len(missing) > 5:
                    preview += f' и ещё {len(missing) - 5}'
                raise EmployeeWorkbookError(
                    'Не найдены сотрудники из файла: ' + preview + '. '
                    'Скачайте свежую версию справочника.'
                )

            changed = []
            for employee in employees:
                old = current[employee['contragent']]
                old_values = (
                    old.get('department') or None,
                    old.get('position') or None,
                    old.get('status') or ALLOWED_STATUSES[0],
                )
                new_values = (
                    employee['department'], employee['position'], employee['status'],
                )
                if old_values != new_values:
                    changed.append((*new_values, employee['contragent']))

            if changed:
                cursor.executemany(
                    '''UPDATE reporting.employees
                       SET department = %s, position = %s, status = %s, updated_at = now()
                       WHERE contragent = %s''',
                    changed,
                )
    return {
        'processed': len(employees),
        'updated': len(changed),
        'unchanged': len(employees) - len(changed),
    }
