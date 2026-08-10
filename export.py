"""Единый экспорт отчётов в Excel.

``build_workbook`` получает уже рассчитанные данные, которые используются на
страницах отчётов, и оформляет их одинаково: закрепляет заголовки, добавляет
фильтры, форматы чисел и дат, а также подбирает читаемую ширину колонок.
"""
from datetime import date, datetime
from io import BytesIO
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill('solid', fgColor='1F4E78')
HEADER_FONT = Font(name='Aptos', size=11, bold=True, color='FFFFFF')
BODY_FONT = Font(name='Aptos', size=10)
ALT_ROW_FILL = PatternFill('solid', fgColor='F4F7FA')
NUMBER_FORMAT = '#,##0.00;[Red](#,##0.00);-'
INTEGER_FORMAT = '#,##0;[Red](#,##0);-'
PERCENT_FORMAT = '0.0\\%'
DATE_FORMAT = 'yyyy-mm-dd'
_INVALID_SHEET_CHARS = re.compile(r'[\\/*?:\[\]]')


def build_workbook(sheets):
    """sheets: [(sheet_name, headers, rows), ...] -> BytesIO(.xlsx)"""
    wb = Workbook()
    wb.remove(wb.active)
    used_sheet_names = set()
    for sheet_index, (name, headers, rows) in enumerate(sheets, start=1):
        ws = wb.create_sheet(_unique_sheet_name(name, used_sheet_names, sheet_index))
        safe_headers = [_cell_safe(header) for header in headers]
        ws.append(safe_headers)
        for row in rows:
            ws.append([_cell_safe(v) for v in row])
        _format_sheet(ws, safe_headers)

    wb.properties.creator = 'Alfa Top Report'
    wb.properties.title = 'Экспорт отчётов Alfa Top Report'
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _cell_safe(v):
    """Приводит сложные значения к строке и блокирует формулы из данных."""
    if isinstance(v, (list, tuple, dict)):
        v = str(v)
    if isinstance(v, str) and v.lstrip().startswith(('=', '+', '-', '@')):
        return "'" + v
    return v


def _unique_sheet_name(name, used, fallback_index):
    base = _INVALID_SHEET_CHARS.sub(' ', str(name or f'Отчёт {fallback_index}')).strip() or f'Отчёт {fallback_index}'
    base = base[:31]
    candidate = base
    suffix = 2
    while candidate.casefold() in used:
        marker = f' ({suffix})'
        candidate = f'{base[:31 - len(marker)]}{marker}'
        suffix += 1
    used.add(candidate.casefold())
    return candidate


def _format_sheet(ws, headers):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'B2' if len(headers) > 1 else 'A2'
    ws.row_dimensions[1].height = 28
    if ws.max_column and ws.max_row:
        ws.auto_filter.ref = f'A1:{get_column_letter(ws.max_column)}{ws.max_row}'

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for row_index in range(2, ws.max_row + 1):
        for cell in ws[row_index]:
            cell.font = BODY_FONT
            cell.alignment = Alignment(
                horizontal='right' if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool) else 'left',
                vertical='center',
            )
            if row_index % 2 == 0:
                cell.fill = ALT_ROW_FILL

    for column_index, header in enumerate(headers, start=1):
        header_text = str(header or '').casefold()
        for row_index in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_index, column=column_index)
            if isinstance(cell.value, (date, datetime)):
                cell.number_format = DATE_FORMAT
            elif isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                if '%' in header_text:
                    cell.number_format = PERCENT_FORMAT
                elif any(token in header_text for token in ('шт', 'чел', 'ед.', 'месяц', 'возраст')):
                    cell.number_format = INTEGER_FORMAT
                else:
                    cell.number_format = NUMBER_FORMAT

        max_length = len(str(header or ''))
        for row_index in range(2, ws.max_row + 1):
            value = ws.cell(row=row_index, column=column_index).value
            if value is None:
                continue
            if isinstance(value, (date, datetime)):
                display_length = 10
            else:
                display_length = len(str(value))
            max_length = max(max_length, display_length)
        max_width = 52 if column_index == 1 else 28
        ws.column_dimensions[get_column_letter(column_index)].width = min(max(max_length + 2, 11), max_width)


def flatten_rows(rows, val_fields):
    """rows: [{'label':, val_field: [..] или число, ...}, ...] -> [[label, *значения], ...].
    Используется свод-таблицами (Свод1/2, UNIT+PL, Dashboard1/2, ФОТ) — каждая строка уже
    содержит один или несколько списков значений по месяцам/сериям плюс, опционально,
    скалярные поля (например endpoint_delta на UNIT+PL)."""
    out = []
    for r in rows:
        row_out = [r.get('label', '')]
        for field in val_fields:
            v = r.get(field)
            if v is None:
                continue
            if isinstance(v, (list, tuple)):
                row_out.extend(v)
            else:
                row_out.append(v)
        out.append(row_out)
    return out
