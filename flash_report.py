"""Flash-отчёт: разбор банковских выписок в предварительную (неаудированную)
картину месяца — до того, как бухгалтер вручную классифицирует их и они попадут
в public."FinancialData" обычным ежемесячным процессом (см. monthly_etl.py).

Классификация не изобретается с нуля: "Комментарии" в public."FinancialData"
совпадает 1-в-1 с "Назначение платежа" из выписки (проверено на реальных июньских
данных), поэтому правила классификации ВЫУЧИВАЮТСЯ эмпирически — сопоставлением
уже разобранных строк выписки с уже классифицированными строками FinancialData
за тот же период по (Дата, Сумма, Контрагент) — см. learn_rules(). Новые
контрагенты/формулировки, которых ещё не было в истории, остаются
"не размечено" и ждут точечной правки бухгалтера (см. flash.classification_rules,
source='manual').

Поддерживаемые форматы выписок (определяются по структуре листа, см. detect_format):
  alfabank      — "_Выписка_<счёт>_<период>.xlsx", шапка + двухстрочный заголовок
  sberbusiness  — "_СберБизнес. Выписка ... счёт <счёт>.xlsx", "плавающие" колонки
  flat28        — плоский формат в один заголовок, 28 колонок (Т-Банк и т.п.)
  svoybank      — "statement_*.xls" — на деле SpreadsheetML XML (Excel 2003 XML),
                  не бинарный xls и не xlsx
"""
import re
import xml.etree.ElementTree as ET
from datetime import date, datetime
from collections import defaultdict

import openpyxl

from db import execute, execute_values, query, query_one

_XML_NS = {'ss': 'urn:schemas-microsoft-com:office:spreadsheet'}

BONUS_LINES = None  # зарезервировано, не используется — flash не размечает бонусы


def _norm_amount(v):
    if v is None or v == '':
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace('\xa0', '').replace(' ', '').replace(',', '.')
    if not s:
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _norm_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    m = re.match(r'^(\d{2})\.(\d{2})\.(\d{4})', s)
    if m:
        try:
            return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            return None
    return None


def _norm_inn(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s and s != '0' else None


def _account_from_filename(filename):
    m = re.search(r'(\d{16,26})', filename)
    return m.group(1) if m else None


# ---------------------------------------------------------------------------
# Формат 1: АльфаБанк — "_Выписка_<счёт>_<период>.xlsx"
# ---------------------------------------------------------------------------

def _parse_alfabank(ws, filename):
    account = _account_from_filename(filename)
    rows = []
    for r in ws.iter_rows(min_row=13, values_only=True):
        op_date = _norm_date(r[0] if len(r) > 0 else None)
        if op_date is None:
            continue
        debet = _norm_amount(r[2] if len(r) > 2 else None)
        kredit = _norm_amount(r[3] if len(r) > 3 else None)
        if debet == 0 and kredit == 0:
            continue
        amount = kredit if kredit else -debet
        rows.append({
            'operation_date': op_date,
            'document_number': str(r[1]) if len(r) > 1 and r[1] is not None else None,
            'amount': amount,
            'counterparty_name': r[4] if len(r) > 4 else None,
            'counterparty_inn': _norm_inn(r[5] if len(r) > 5 else None),
            'purpose_text': r[10] if len(r) > 10 else None,
            'account_number': account,
        })
    return rows


# ---------------------------------------------------------------------------
# Формат 2: СберБизнес — "плавающие" колонки, Счёт-Дебет/Счёт-Кредит через \n
# ---------------------------------------------------------------------------

def _split_packed_account(cell):
    """"Счет\\nИНН\\nНаименование" -> (account, inn, name)."""
    if not cell:
        return None, None, None
    parts = str(cell).split('\n')
    parts = [p.strip() for p in parts if p.strip()]
    account = parts[0] if len(parts) > 0 else None
    inn = _norm_inn(parts[1]) if len(parts) > 1 else None
    name = parts[2] if len(parts) > 2 else None
    return account, inn, name


def _parse_sberbusiness(ws, filename):
    account = _account_from_filename(filename)
    rows = []
    for r in ws.iter_rows(min_row=12, values_only=True):
        op_date = _norm_date(r[1] if len(r) > 1 else None)
        if op_date is None:
            continue
        debet_side = r[4] if len(r) > 4 else None
        kredit_side = r[8] if len(r) > 8 else None
        sum_debet = _norm_amount(r[9] if len(r) > 9 else None)
        sum_kredit = _norm_amount(r[13] if len(r) > 13 else None)
        if sum_debet == 0 and sum_kredit == 0:
            continue
        if sum_debet:
            amount = -sum_debet
            cp_account, cp_inn, cp_name = _split_packed_account(kredit_side)
        else:
            amount = sum_kredit
            cp_account, cp_inn, cp_name = _split_packed_account(debet_side)
        rows.append({
            'operation_date': op_date,
            'document_number': str(r[14]) if len(r) > 14 and r[14] is not None else None,
            'amount': amount,
            'counterparty_name': cp_name,
            'counterparty_inn': cp_inn,
            'purpose_text': r[20] if len(r) > 20 else None,
            'account_number': account,
        })
    return rows


# ---------------------------------------------------------------------------
# Формат 3: плоский, 28 колонок (Т-Банк и аналогичные)
# ---------------------------------------------------------------------------

def _parse_flat28(ws, filename):
    rows = []
    for r in ws.iter_rows(min_row=11, values_only=True):
        op_date = _norm_date(r[2] if len(r) > 2 else None)
        if op_date is None:
            continue
        debet = _norm_amount(r[26] if len(r) > 26 else None)
        kredit = _norm_amount(r[27] if len(r) > 27 else None)
        if debet == 0 and kredit == 0:
            continue
        amount = kredit if kredit else -debet
        rows.append({
            'operation_date': op_date,
            'document_number': str(r[3]) if len(r) > 3 and r[3] is not None else None,
            'amount': amount,
            'counterparty_name': r[24] if len(r) > 24 else None,
            'counterparty_inn': _norm_inn(r[23] if len(r) > 23 else None),
            'purpose_text': (r[8] if len(r) > 8 else None) or (r[7] if len(r) > 7 else None),
            'account_number': r[0] if len(r) > 0 else _account_from_filename(filename),
        })
    return rows


# ---------------------------------------------------------------------------
# Формат 4: Свой Банк — SpreadsheetML XML (Excel 2003 XML) под .xls
# ---------------------------------------------------------------------------

def _xml_row_cells(row_elem):
    """Учитывает ss:Index (пропущенные ячейки в XML не идут подряд)."""
    cells = {}
    idx = 0
    for c in row_elem.findall('ss:Cell', _XML_NS):
        ss_index = c.get('{urn:schemas-microsoft-com:office:spreadsheet}Index')
        if ss_index:
            idx = int(ss_index) - 1
        data = c.find('ss:Data', _XML_NS)
        cells[idx] = data.text if data is not None else None
        idx += 1
    if not cells:
        return []
    width = max(cells) + 1
    return [cells.get(i) for i in range(width)]


def _parse_svoybank_xml(path_or_bytes, filename):
    tree = ET.parse(path_or_bytes)
    root = tree.getroot()
    ws = root.find('ss:Worksheet', _XML_NS)
    table = ws.find('ss:Table', _XML_NS)
    all_rows = table.findall('ss:Row', _XML_NS)

    account = None
    if all_rows:
        first = _xml_row_cells(all_rows[0])
        if first:
            m = re.search(r'(\d{16,26})', str(first[0] or ''))
            account = m.group(1) if m else None
    if account is None:
        account = _account_from_filename(filename)

    rows = []
    for row_elem in all_rows:
        r = _xml_row_cells(row_elem)
        if len(r) < 10:
            continue
        op_date = _norm_date(r[1])
        if op_date is None:
            continue
        oborot_dt = _norm_amount(r[8])
        oborot_kt = _norm_amount(r[9])
        if oborot_dt == 0 and oborot_kt == 0:
            continue
        amount = oborot_kt if oborot_kt else -oborot_dt
        rows.append({
            'operation_date': op_date,
            'document_number': r[0],
            'amount': amount,
            'counterparty_name': r[2],
            'counterparty_inn': _norm_inn(r[3]),
            'purpose_text': r[10] if len(r) > 10 else None,
            'account_number': account,
        })
    return rows


# ---------------------------------------------------------------------------
# Определение формата + точка входа
# ---------------------------------------------------------------------------

def detect_and_parse(file_obj, filename):
    """Возвращает (bank_format, rows) — rows это список нормализованных словарей
    (operation_date, document_number, amount, counterparty_name, counterparty_inn,
    purpose_text, account_number), ещё без классификации."""
    lower = filename.lower()
    if lower.endswith('.xls') and not lower.endswith('.xlsx'):
        # Проверено на реальном файле: .xls тут — на деле SpreadsheetML XML,
        # не бинарный формат (Excel так экспортирует "Свой Банк" и похожие).
        rows = _parse_svoybank_xml(file_obj, filename)
        return 'svoybank', rows

    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row1 = [ws.cell(row=1, column=c).value for c in range(1, min(ws.max_column, 5) + 1)]
    header_row10 = [ws.cell(row=10, column=c).value for c in range(1, min(ws.max_column, 5) + 1)]
    header_row11 = [ws.cell(row=11, column=c).value for c in range(1, min(ws.max_column, 5) + 1)]

    if header_row1 and header_row1[0] == 'Выписка по счёту':
        return 'alfabank', _parse_alfabank(ws, filename)
    if header_row10 and header_row10[0] == 'Номер счёта' and ws.max_column >= 27:
        return 'flat28', _parse_flat28(ws, filename)
    if header_row11 and 'Дебет' in (header_row11[4] if len(header_row11) > 4 else ''):
        return 'sberbusiness', _parse_sberbusiness(ws, filename)
    # СберБизнес-заголовок иногда съезжает на строку — проверим окрестность.
    for probe in range(9, 13):
        row = [ws.cell(row=probe, column=c).value for c in range(1, 6)]
        if row and row[4] == 'Дебет':
            return 'sberbusiness', _parse_sberbusiness(ws, filename)

    raise ValueError(
        f'Не удалось определить формат выписки для файла "{filename}" — '
        'нужна доработка парсера под новый формат банка.'
    )


# ---------------------------------------------------------------------------
# Классификация: выучивание правил из уже классифицированной истории + применение
# ---------------------------------------------------------------------------

RULE_CONFIDENCE_THRESHOLD = 0.85
# Один контрагент (по ИНН) или один и тот же текст назначения платежа
# (purpose_contains) на практике сплошь и рядом относится к РАЗНЫМ строкам
# отчёта и проектам — проверено на реальных данных: "Рубанков Иван
# Владимирович" / текст "Оплата самозанятому за услуги по договору № 21/КА"
# за май-июнь распадается на 4 разные "Строка отчета" и 11 разных "Проект"
# (один платёж полевому агенту распределяется по проектам его загрузки —
# сигнала для этого распределения нет ни в ИНН, ни в тексте платежа, только
# в отдельной таблице драйверов, которой у Flash нет). Обобщающее правило
# создаём ТОЛЬКО когда один вариант набирает свою долю от общей суммы
# сопоставленных операций >= порога — иначе платёж остаётся неразмеченным
# для бухгалтера, вместо угадывания наугад.


def learn_rules(period, created_by='system'):
    """Двухшаговый процесс:
    1) Каждую операцию Flash за period сопоставляет её СОБСТВЕННОЙ строке
       public."FinancialData" по (Дата, round(Сумма,2)) — при нескольких
       кандидатах предпочитает совпадение "Комментарии"/purpose_text, а если
       текстового совпадения нет и кандидаты расходятся по "Строка отчета",
       НЕ гадает (оставляет операцию неразмеченной) — раньше здесь молча
       брался первый попавшийся кандидат, что путало напрямую несвязанные
       проводки, совпавшие по (Дата, Сумма) случайно.
    2) Строит ОБОБЩАЮЩЕЕ правило (для будущих операций без точного совпадения
       по сумме — например, разбитые по проектам платежи агентам) только
       если по всем найденным в шаге 1 операциям того же ИНН/текста платежа
       один вариант "Строка отчета" набирает >= RULE_CONFIDENCE_THRESHOLD от
       суммы; "Проект" в правило попадает отдельным, более строгим условием
       (тот же порог, но уже внутри доминирующей строки).
    Возвращает число новых/обновлённых обобщающих правил."""
    fd_rows = query(
        '''SELECT "Дата", "Сумма", "Контрагент", "Комментарии", "Признак", "Категория",
                  "Статья", "Проект", "Контрагент_report", "Строка отчета"
           FROM public."FinancialData"
           WHERE "Период" = %s AND "п_ф" = 'факт' AND "Распределение" = 'до распределения' ''',
        (period,)
    )
    by_key = defaultdict(list)
    for r in fd_rows:
        key = (r['Дата'], round(float(r['Сумма'] or 0), 2))
        by_key[key].append(r)

    txns = query(
        '''SELECT id, operation_date, amount, counterparty_name, counterparty_inn, purpose_text
           FROM flash.transactions WHERE operation_date >= %s AND operation_date < (%s::date + interval '1 month')''',
        (period, period)
    )

    # Шаг 1 — точное сопоставление каждой операции со своей строкой FinancialData.
    txn_matches = []
    for t in txns:
        key = (t['operation_date'], round(float(t['amount']), 2))
        candidates = by_key.get(key)
        if not candidates:
            continue
        match = None
        if len(candidates) == 1:
            match = candidates[0]
        else:
            if t['purpose_text']:
                for c in candidates:
                    if c['Комментарии'] and t['purpose_text'][:30] in (c['Комментарии'] or ''):
                        match = c
                        break
            if match is None:
                lines = {c['Строка отчета'] for c in candidates}
                if len(lines) == 1:
                    match = candidates[0]
        if match is not None:
            txn_matches.append((t, match))

    # Существующие правила грузим одним запросом и матчим/апдейтим в памяти —
    # раньше здесь было 3 SQL-запроса НА КАЖДУЮ строку выписки (существующее
    # правило? апдейт/инсерт правила; апдейт транзакции), что на удалённой
    # Render Postgres для ~3300 строк заняло больше 100 минут. Батчим через
    # execute_values — единицы round-trip'ов вместо тысяч.
    existing_rules = query('SELECT id, match_type, match_value FROM flash.classification_rules')
    existing_by_key = {(r['match_type'], r['match_value']): r['id'] for r in existing_rules}

    # Шаг 1.5 — применяем точный, проверенный результат к самим этим операциям
    # (это реальная истина именно для них, не обобщение).
    txn_updates = [
        (fd['Признак'], fd['Категория'], fd['Статья'], fd['Проект'], fd['Контрагент_report'],
         fd['Строка отчета'], 'learned', t['id'])
        for t, fd in txn_matches
    ]

    # Шаг 2 — обобщающие правила по ИНН / тексту платежа, только при уверенном большинстве.
    inn_records = defaultdict(list)
    purpose_records = defaultdict(list)
    for t, fd in txn_matches:
        amt = abs(float(t['amount']))
        line = (fd['Признак'], fd['Категория'], fd['Статья'], fd['Контрагент_report'], fd['Строка отчета'])
        if t['counterparty_inn']:
            inn_records[t['counterparty_inn']].append((amt, line, fd['Проект']))
        elif t['purpose_text']:
            mv = t['purpose_text'][:60].strip()
            purpose_records[mv].append((amt, line, fd['Проект']))

    rule_inserts = {}
    rule_updates = {}

    def _build_rule(match_type, match_value, records):
        line_totals = defaultdict(float)
        total = 0.0
        for amt, line, _proekt in records:
            line_totals[line] += amt
            total += amt
        if total <= 0:
            return
        dominant_line, dominant_amt = max(line_totals.items(), key=lambda kv: kv[1])
        if dominant_amt / total < RULE_CONFIDENCE_THRESHOLD:
            return

        proj_totals = defaultdict(float)
        proj_total = 0.0
        for amt, line, proekt in records:
            if line == dominant_line:
                proj_totals[proekt] += amt
                proj_total += amt
        proekt_val = None
        if proj_total > 0:
            dom_proj, dom_proj_amt = max(proj_totals.items(), key=lambda kv: kv[1])
            if dom_proj_amt / proj_total >= RULE_CONFIDENCE_THRESHOLD:
                proekt_val = dom_proj

        priznak, kategoria, statya, kontragent, stroka = dominant_line
        fields = (priznak, kategoria, statya, proekt_val, kontragent, stroka)
        rule_key = (match_type, match_value)
        rid = existing_by_key.get(rule_key)
        if rid:
            rule_updates[rid] = fields + (rid,)
        else:
            rule_inserts[rule_key] = (match_type, match_value) + fields + ('learned', created_by)

    for inn, records in inn_records.items():
        _build_rule('inn', inn, records)
    for mv, records in purpose_records.items():
        _build_rule('purpose_contains', mv, records)

    learned = len(rule_inserts)
    if rule_inserts:
        execute_values(
            '''INSERT INTO flash.classification_rules
               (match_type, match_value, "Признак", "Категория", "Статья", "Проект", "Контрагент_report",
                "Строка отчета", source, created_by)
               VALUES %s''',
            list(rule_inserts.values())
        )
    if rule_updates:
        execute_values(
            '''UPDATE flash.classification_rules AS r SET
                 "Признак" = v.priznak, "Категория" = v.kategoria, "Статья" = v.statya,
                 "Проект" = v.proekt, "Контрагент_report" = v.kontragent, "Строка отчета" = v.stroka
               FROM (VALUES %s) AS v(priznak, kategoria, statya, proekt, kontragent, stroka, id)
               WHERE r.id = v.id''',
            list(rule_updates.values())
        )
    if txn_updates:
        execute_values(
            '''UPDATE flash.transactions AS t SET
                 "Признак" = v.priznak, "Категория" = v.kategoria, "Статья" = v.statya,
                 "Проект" = v.proekt, "Контрагент_report" = v.kontragent, "Строка отчета" = v.stroka,
                 classification_source = v.source
               FROM (VALUES %s) AS v(priznak, kategoria, statya, proekt, kontragent, stroka, source, id)
               WHERE t.id = v.id''',
            txn_updates
        )
    return learned


def _load_classification_rules():
    """Все правила одним запросом — для построчной классификации в цикле
    (import_statement) без N+1 запросов на удалённую БД. Возвращает
    (inn_rules: {инн: правило}, purpose_rules: [правило, ...] отсортированные
    по priority DESC, длине match_value DESC — так же, как раньше это делал
    ORDER BY в classify_transaction)."""
    rules = query(
        '''SELECT id, match_type, match_value, "Признак", "Категория", "Статья", "Проект", "Контрагент_report",
                  "Строка отчета"
           FROM flash.classification_rules ORDER BY priority DESC, length(match_value) DESC'''
    )
    inn_rules = {}
    purpose_rules = []
    for r in rules:
        if r['match_type'] == 'inn':
            inn_rules.setdefault(r['match_value'], r)
        else:
            purpose_rules.append(r)
    return inn_rules, purpose_rules


def _load_wallet_aliases():
    rows = query('SELECT account_number, wallet FROM flash.wallet_aliases')
    return {r['account_number']: r['wallet'] for r in rows}


def _match_rule(txn, inn_rules, purpose_rules):
    """Сопоставление в памяти против уже загруженных правил (см.
    _load_classification_rules) — сначала по ИНН (надёжнее), потом по
    вхождению начала правила в текст назначения платежа. Возвращает
    (правило, rule_id) либо (None, None)."""
    if txn['counterparty_inn'] and txn['counterparty_inn'] in inn_rules:
        rule = inn_rules[txn['counterparty_inn']]
        return rule, rule['id']
    if txn['purpose_text']:
        for r in purpose_rules:
            if r['match_value'] and r['match_value'] in txn['purpose_text']:
                return r, r['id']
    return None, None


def reclassify_unmatched(period):
    """Повторно применяет ТЕКУЩИЕ flash.classification_rules к уже
    загруженным неразмеченным операциям месяца — для периодов, у которых ещё
    нет своей "факт"-FinancialData (learn_rules() тогда ничего не находит,
    т.к. сверять не с чем), но уже есть правила, выученные по предыдущим
    закрытым месяцам. Тот же батч-путь _load_classification_rules/_match_rule,
    что использует import_statement(), только по уже сохранённым в базе
    операциям, а не по свежераспарсенному файлу. Возвращает число размеченных."""
    inn_rules, purpose_rules = _load_classification_rules()
    txns = query(
        '''SELECT id, counterparty_inn, purpose_text FROM flash.transactions
           WHERE date_trunc('month', operation_date) = %s AND classification_source = 'unmatched' ''',
        (period,)
    )
    updates = []
    for t in txns:
        rule, _rule_id = _match_rule(t, inn_rules, purpose_rules)
        if rule:
            updates.append((
                rule['Признак'], rule['Категория'], rule['Статья'], rule['Проект'],
                rule['Контрагент_report'], rule['Строка отчета'], 'rule', t['id']
            ))
    if updates:
        execute_values(
            '''UPDATE flash.transactions AS t SET
                 "Признак" = v.priznak, "Категория" = v.kategoria, "Статья" = v.statya,
                 "Проект" = v.proekt, "Контрагент_report" = v.kontragent, "Строка отчета" = v.stroka,
                 classification_source = v.source
               FROM (VALUES %s) AS v(priznak, kategoria, statya, proekt, kontragent, stroka, source, id)
               WHERE t.id = v.id''',
            updates
        )
    return len(updates)


def classify_transaction(txn):
    """Применяет flash.classification_rules к одной операции — для точечной
    переклассификации вне массового импорта (сам import_statement использует
    более быстрый батч-путь через _load_classification_rules/_match_rule,
    чтобы не делать 2+ запроса на каждую из тысяч строк выписки).
    Возвращает (fields_dict, source) — source='rule'/'unmatched'."""
    inn_rules, purpose_rules = _load_classification_rules()
    rule, rule_id = _match_rule(txn, inn_rules, purpose_rules)
    if rule:
        execute('UPDATE flash.classification_rules SET hits = hits + 1 WHERE id = %s', (rule_id,))
        return rule, 'rule'
    return None, 'unmatched'


def learn_wallet_aliases(period, created_by='system'):
    """Определяет "Кошелек" для каждого account_number в flash.transactions —
    по большинству голосов среди совпавших через learn_rules() строк
    public."FinancialData" за тот же период (см. matched_financialdata_id — но
    проще: сопоставляем ещё раз по (Дата, Сумма) и берём "Кошелек" оттуда)."""
    accounts = query(
        'SELECT DISTINCT account_number FROM flash.transactions WHERE account_number IS NOT NULL'
    )
    fd_rows = query(
        '''SELECT "Дата", "Сумма", "Кошелек", "Тип Кошелька" FROM public."FinancialData"
           WHERE "Период" = %s AND "п_ф" = 'факт' AND "Распределение" = 'до распределения' ''',
        (period,)
    )
    by_key = defaultdict(list)
    for r in fd_rows:
        by_key[(r['Дата'], round(float(r['Сумма'] or 0), 2))].append(r)

    learned = 0
    for acc in accounts:
        account_number = acc['account_number']
        txns = query(
            'SELECT operation_date, amount FROM flash.transactions WHERE account_number = %s',
            (account_number,)
        )
        votes = defaultdict(int)
        wallet_type_by_wallet = {}
        for t in txns:
            key = (t['operation_date'], round(float(t['amount']), 2))
            for r in by_key.get(key, []):
                if r['Кошелек']:
                    votes[r['Кошелек']] += 1
                    wallet_type_by_wallet[r['Кошелек']] = r['Тип Кошелька']
        if not votes:
            continue
        wallet = max(votes, key=votes.get)
        existing = query_one('SELECT id FROM flash.wallet_aliases WHERE account_number = %s', (account_number,))
        if existing:
            execute('UPDATE flash.wallet_aliases SET wallet = %s, wallet_type = %s WHERE id = %s',
                    (wallet, wallet_type_by_wallet.get(wallet), existing['id']))
        else:
            execute(
                '''INSERT INTO flash.wallet_aliases (account_number, wallet, wallet_type, source, created_by)
                   VALUES (%s, %s, %s, 'learned', %s)''',
                (account_number, wallet, wallet_type_by_wallet.get(wallet), created_by)
            )
            learned += 1
        execute('UPDATE flash.transactions SET wallet = %s WHERE account_number = %s', (wallet, account_number))
    return learned


def wallet_reconciliation(period):
    """Сверка кассового движения по Flash с уже известными остатками из
    "Сверка счетов" (wallet_report) — по той же канонической привязке
    account_number -> "Кошелек", что уже выучена в flash.wallet_aliases
    (см. learn_wallet_aliases) и показана пользователю в колонке "Кошелёк".

    Один физический счёт иногда фактически финансирует больше одного
    бухгалтерского "Кошелька" (проверено на реальных данных) — но надёжно
    различить это по (Дата, Сумма) не получается: на некрупных круглых суммах
    попадается заметная доля случайных совпадений с совершенно другими
    кошельками (до ~10% операций счёта в проверке), из-за чего наивное
    "просуммировать остатки всех когда-либо совпавших кошельков" завышает
    результат чужими остатками. Поэтому берём только доминирующий кошелёк
    счёта (как и остальной Flash) — известное упрощение, но оно не путает
    счета между собой.

    Возвращает по каждому кошельку: opening (остаток на начало периода из
    Сверка счетов), flash_turnover (оборот по Flash за period), calculated
    (opening + flash_turnover), known_after (если за period в Сверка счетов
    уже введён факт вручную — иначе None), discrepancy (known_after -
    calculated, если известно)."""
    import wallet_report as wr

    turnover_rows = query(
        '''SELECT wallet, SUM(amount) AS amt, count(*) AS cnt FROM flash.transactions
           WHERE date_trunc('month', operation_date) = %s AND wallet IS NOT NULL
           GROUP BY 1 ORDER BY 1''',
        (period,)
    )

    rows = []
    for r in turnover_rows:
        wname = r['wallet']
        detail = wr.wallet_detail(wname, year='all')
        opening = 0.0
        known_after = None
        if detail:
            prior = [row for row in detail['rows'] if row['period'] < period]
            if prior:
                opening = prior[-1]['balance']
            same = next((row for row in detail['rows'] if row['period'] == period), None)
            if same and same['entered'] is not None:
                known_after = same['entered']

        flash_amt = float(r['amt'])
        calculated = opening + flash_amt
        rows.append({
            'wallet': wname,
            'opening': opening,
            'flash_turnover': flash_amt,
            'cnt': r['cnt'],
            'calculated': calculated,
            'known_after': known_after,
            'discrepancy': (known_after - calculated) if known_after is not None else None,
        })
    return rows


def import_statement(file_obj, filename, username):
    """Разбирает один файл выписки, классифицирует построчно по уже выученным
    правилам и сохраняет в flash.transactions. ON CONFLICT — по содержимому
    операции (bank_format, account_number, document_number, дата, сумма, ИНН,
    назначение платежа), не по порядку строк в файле: выписки грузятся
    еженедельно/ежедневно растущим диапазоном дат, и повторная загрузка того
    же (или расширенного) файла добавляет только новые строки, уже
    размеченные не трогает — даже если банк отдал строки в другом порядке.
    Дедуп-индекс — idx_flash_transactions_dedup (см. schema.sql), с COALESCE
    на ИНН/назначение платежа: у операций от физлиц-должников counterparty_inn
    часто NULL, а в Postgres NULL никогда не равен NULL, так что обычный
    UNIQUE(...) такие строки не считал бы дубликатами — было реально
    проверено на живых данных: повторная загрузка того же файла задвоила
    ровно те 2 строки, где ИНН плательщика отсутствовал.

    Правила/алиасы кошельков грузятся один раз на файл и классификация идёт
    в памяти (не по запросу на строку) — на файле в тысячи операций на
    удалённой Render Postgres это разница между минутами и секундами.
    Возвращает {bank_format, total, matched}."""
    bank_format, rows = detect_and_parse(file_obj, filename)
    if not rows:
        raise ValueError(f'В файле "{filename}" не найдено ни одной операции — проверь формат.')

    inn_rules, purpose_rules = _load_classification_rules()
    wallet_by_account = _load_wallet_aliases()

    matched = 0
    hit_rule_ids = []
    values = []
    for r in rows:
        rule, rule_id = _match_rule(r, inn_rules, purpose_rules)
        if rule:
            source = 'rule'
            matched += 1
            hit_rule_ids.append(rule_id)
            fields = (rule['Признак'], rule['Категория'], rule['Статья'], rule['Проект'], rule['Контрагент_report'],
                      rule['Строка отчета'])
        else:
            source = 'unmatched'
            fields = (None, None, None, None, None, None)

        wallet = wallet_by_account.get(r['account_number'])
        values.append((
            filename, bank_format, r['account_number'], wallet, r['operation_date'], r['document_number'], r['amount'],
            r['counterparty_name'], r['counterparty_inn'], r['purpose_text'],
        ) + fields + (source, username))

    execute_values(
        '''INSERT INTO flash.transactions
           (source_file, bank_format, account_number, wallet, operation_date, document_number, amount,
            counterparty_name, counterparty_inn, purpose_text,
            "Признак", "Категория", "Статья", "Проект", "Контрагент_report", "Строка отчета",
            classification_source, imported_by)
           VALUES %s
           ON CONFLICT (bank_format, account_number, document_number, operation_date, amount,
                        COALESCE(counterparty_inn, ''), COALESCE(purpose_text, ''))
           DO NOTHING''',
        values
    )
    if hit_rule_ids:
        execute('UPDATE flash.classification_rules SET hits = hits + 1 WHERE id = ANY(%s)', (hit_rule_ids,))

    return {'bank_format': bank_format, 'total': len(rows), 'matched': matched}


def _effective_rows_sql():
    """SQL (без WHERE/ORDER) операций Flash с учётом ручного разбиения по
    проектам (flash.transaction_splits, см. set_transaction_splits): у
    разбитых операций вместо одной строки исходной транзакции подставляются
    её части (каждая со своей суммой/классификацией), а сама исходная
    транзакция (classification_source='split') из результата исключается —
    иначе сумма задвоилась бы. split_id IS NULL — обычная (неразбитая)
    операция, iначе — конкретная часть разбивки."""
    return '''
        SELECT t.id, NULL::int AS split_id, t.operation_date, t.amount, t.counterparty_name,
               t.counterparty_inn, t.purpose_text, t.wallet, t.bank_format, t.source_file,
               t."Признак", t."Категория", t."Статья", t."Проект", t."Контрагент_report",
               t."Строка отчета", t.classification_source
        FROM flash.transactions t
        WHERE t.classification_source != 'split'
        UNION ALL
        SELECT t.id, s.id AS split_id, t.operation_date, s.amount, t.counterparty_name,
               t.counterparty_inn, t.purpose_text, t.wallet, t.bank_format, t.source_file,
               s."Признак", s."Категория", s."Статья", s."Проект", s."Контрагент_report",
               s."Строка отчета", 'manual' AS classification_source
        FROM flash.transaction_splits s JOIN flash.transactions t ON t.id = s.transaction_id
    '''


def summary(period=None):
    """Свод по Flash: обороты по месяцам (как Свод1), разбивка размечено/не
    размечено по сумме — предварительная картина, помечается на странице как
    неаудированная. classification_source: 'rule'/'learned'/'manual' — это всё
    успешная классификация (по правилу, автоматом при обучении, вручную
    бухгалтером); 'unmatched' — единственное реально неразмеченное состояние."""
    where = ''
    params = ()
    if period:
        where = "WHERE date_trunc('month', operation_date) = %s"
        params = (period,)
    rows = query(
        f'''SELECT date_trunc('month', operation_date)::date AS period, "Статья", "Категория", "Признак", "Проект",
                   classification_source, SUM(amount) AS amount, count(*) AS cnt
           FROM ({_effective_rows_sql()}) AS ft {where}
           GROUP BY 1, 2, 3, 4, 5, 6 ORDER BY 1''',
        params
    )
    total_matched = sum(float(r['amount']) for r in rows if r['classification_source'] != 'unmatched')
    total_unmatched = sum(float(r['amount']) for r in rows if r['classification_source'] == 'unmatched')
    cnt_matched = sum(r['cnt'] for r in rows if r['classification_source'] != 'unmatched')
    cnt_unmatched = sum(r['cnt'] for r in rows if r['classification_source'] == 'unmatched')
    return {
        'rows': rows, 'total_matched': total_matched, 'total_unmatched': total_unmatched,
        'cnt_matched': cnt_matched, 'cnt_unmatched': cnt_unmatched,
    }


def month_breakdown(period):
    """Разбивка одного месяца по тем же секциям, что и Свод1 (Выручка/
    Переменные/Постоянные/Инвестиции в портфели/Финансирование/Бонусы —
    pl_report.REVENUE_LINES и т.д.), только размеченные операции
    (classification_source != 'unmatched'), на сырых данных выписки, без
    ожидания ручной классификации бухгалтера. profit/gm/net_profit считаются
    по тем же формулам, что pl_report.svod1() — кассовая прикидка того же
    П&Л по банковским движениям."""
    import pl_report as pr
    rows = query(
        f'''SELECT "Статья", "Строка отчета", SUM(amount) AS amount, count(*) AS cnt
           FROM ({_effective_rows_sql()}) AS ft
           WHERE date_trunc('month', operation_date) = %s AND classification_source != 'unmatched'
           GROUP BY 1, 2 ORDER BY 3 ASC''',
        (period,)
    )
    buckets = {'revenue': [], 'variable': [], 'fixed': [], 'investment': [], 'financing': [], 'bonus': [], 'other': []}
    totals = {k: 0.0 for k in buckets}
    for r in rows:
        line = r['Строка отчета']
        amt = float(r['amount'])
        if line in pr.REVENUE_LINES:
            key = 'revenue'
        elif line in pr.VARIABLE_LINES:
            key = 'variable'
        elif line in pr.FIXED_LINES:
            key = 'fixed'
        elif line == pr.INVESTMENT_LINE:
            key = 'investment'
        elif line in pr.FINANCING_LINES:
            key = 'financing'
        elif line in pr.BONUS_LINES:
            key = 'bonus'
        else:
            key = 'other'
        buckets[key].append(r)
        totals[key] += amt
    gm = totals['revenue'] + totals['variable']
    profit = gm + totals['fixed']
    net_profit = profit + totals['bonus'] + totals['investment'] + totals['financing']
    return {'buckets': buckets, 'totals': totals, 'gm': gm, 'profit': profit, 'net_profit': net_profit}


def by_project(period):
    """Разбивка размеченных операций месяца по "Проект" (только операции, у
    которых он вообще определён — Проект часто неизвестен, см.
    RULE_CONFIDENCE_THRESHOLD и set_transaction_splits) — для быстрого
    обзора, какие проекты уже видны во Flash за месяц."""
    rows = query(
        f'''SELECT "Проект", SUM(amount) AS amount, count(*) AS cnt
           FROM ({_effective_rows_sql()}) AS ft
           WHERE date_trunc('month', operation_date) = %s AND classification_source != 'unmatched'
                 AND "Проект" IS NOT NULL
           GROUP BY 1 ORDER BY 2 DESC''',
        (period,)
    )
    without_project = query(
        f'''SELECT SUM(amount) AS amount, count(*) AS cnt
           FROM ({_effective_rows_sql()}) AS ft
           WHERE date_trunc('month', operation_date) = %s AND classification_source != 'unmatched'
                 AND "Проект" IS NULL''',
        (period,)
    )[0]
    return {
        'rows': rows,
        'without_project_amount': float(without_project['amount'] or 0),
        'without_project_cnt': without_project['cnt'],
    }


def get_unmatched(period=None, limit=500):
    where = ''
    params = ()
    if period:
        where = "WHERE date_trunc('month', operation_date) = %s AND"
        params = (period,)
    else:
        where = 'WHERE'
    return query(
        f'''SELECT id, operation_date, amount, counterparty_name, counterparty_inn, purpose_text, wallet
           FROM flash.transactions {where} classification_source = 'unmatched'
           ORDER BY abs(amount) DESC LIMIT %s''',
        params + (limit,)
    )


def get_transactions(period, statya=None, stroka=None):
    """Список операций за период — для drill-down по строке отчёта
    (statya/stroka) с странице /flash: и уже размеченные, и нет, чтобы можно
    было провалиться в любую агрегированную строку и отредактировать любую
    отдельную операцию, а не только явно "неразмеченные". split_id — часть
    ручного разбиения (см. set_transaction_splits) или NULL для обычной
    операции; id — всегда id самой транзакции (для разбитых операций
    указывает на исходную flash.transactions, split_id — на конкретную часть)."""
    where = ["date_trunc('month', operation_date) = %s"]
    params = [period]
    if statya is not None:
        where.append('"Статья" = %s')
        params.append(statya)
    if stroka is not None:
        where.append('"Строка отчета" = %s')
        params.append(stroka)
    return query(
        f'''SELECT id, split_id, operation_date, amount, counterparty_name, counterparty_inn, purpose_text, wallet,
                   "Признак", "Категория", "Статья", "Проект", "Контрагент_report", "Строка отчета",
                   classification_source
           FROM ({_effective_rows_sql()}) AS ft WHERE {" AND ".join(where)}
           ORDER BY abs(amount) DESC''',
        tuple(params)
    )


def export_for_review(period):
    """Полный дамп операций месяца (размеченные и нет, все поля) — для ручной
    проверки бухгалтером в Excel, в отличие от export_for_load(), который
    отдаёт только размеченные строки в узком формате листа "загрузка"."""
    rows = query(
        '''SELECT operation_date, amount, counterparty_name, counterparty_inn, purpose_text, wallet,
                  bank_format, source_file, "Признак", "Категория", "Статья", "Проект",
                  "Контрагент_report", "Строка отчета", classification_source
           FROM flash.transactions WHERE date_trunc('month', operation_date) = %s
           ORDER BY operation_date, id''',
        (period,)
    )
    headers = [
        'Дата', 'Сумма', 'Контрагент (банк)', 'ИНН', 'Назначение платежа', 'Кошелёк',
        'Формат выписки', 'Файл', 'Признак', 'Категория', 'Статья', 'Проект',
        'Контрагент (отчёт)', 'Строка отчета', 'Источник разметки',
    ]
    out_rows = [[
        r['operation_date'], float(r['amount']), r['counterparty_name'], r['counterparty_inn'],
        r['purpose_text'], r['wallet'], r['bank_format'], r['source_file'],
        r['Признак'], r['Категория'], r['Статья'], r['Проект'],
        r['Контрагент_report'], r['Строка отчета'], r['classification_source'],
    ] for r in rows]
    return headers, out_rows


def set_manual_classification(txn_id, fields, username):
    """Ручная правка одной операции бухгалтером — заодно заводит/обновляет
    правило по ИНН (или по назначению платежа, если ИНН нет), чтобы это же
    правило подхватило будущие загрузки такой же операции.

    НЕ применяет правило задним числом к другим уже загруженным
    неразмеченным операциям того же контрагента/текста — попробовали так
    раньше и получили реальную порчу данных: один контрагент (тем более один
    и тот же текст назначения платежа, например "Оплата самозанятому...")
    сплошь и рядом относится к разным строкам отчёта и проектам (проверено:
    у одного самозанятого агента один и тот же текст в мае-июне разошёлся на
    4 разные "Строка отчета"), так что решение бухгалтера по ОДНОЙ операции
    не гарантированно верно для остальных — каждую нужно смотреть отдельно
    через drill-down. Массовое обучение правил с проверкой статистической
    уверенности (>= RULE_CONFIDENCE_THRESHOLD доли по сумме) — задача
    learn_rules(), не этой функции.

    fields: Признак, Категория, Статья, Проект, Контрагент_report, Строка
    отчета (последнее — из фиксированного списка pl_report.ALL_LINES,
    определяет, куда строка попадёт в Выручка/Переменные/Постоянные — см.
    month_breakdown)."""
    txn = query_one('SELECT counterparty_inn, purpose_text FROM flash.transactions WHERE id = %s', (txn_id,))
    if txn is None:
        raise ValueError('Операция не найдена')

    values = (fields.get('Признак'), fields.get('Категория'), fields.get('Статья'), fields.get('Проект'),
              fields.get('Контрагент_report'), fields.get('Строка отчета'))

    execute(
        '''UPDATE flash.transactions SET "Признак"=%s, "Категория"=%s, "Статья"=%s, "Проект"=%s,
           "Контрагент_report"=%s, "Строка отчета"=%s, classification_source='manual' WHERE id = %s''',
        values + (txn_id,)
    )

    if txn['counterparty_inn']:
        match_type, match_value = 'inn', txn['counterparty_inn']
    elif txn['purpose_text']:
        match_type, match_value = 'purpose_contains', txn['purpose_text'][:60].strip()
    else:
        return 0

    existing = query_one(
        'SELECT id FROM flash.classification_rules WHERE match_type = %s AND match_value = %s',
        (match_type, match_value)
    )
    if existing:
        execute(
            '''UPDATE flash.classification_rules
               SET "Признак"=%s, "Категория"=%s, "Статья"=%s, "Проект"=%s, "Контрагент_report"=%s,
                   "Строка отчета"=%s, source='manual'
               WHERE id = %s''',
            values + (existing['id'],)
        )
    else:
        execute(
            '''INSERT INTO flash.classification_rules
               (match_type, match_value, "Признак", "Категория", "Статья", "Проект", "Контрагент_report",
                "Строка отчета", priority, source, created_by)
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 10, 'manual', %s)''',
            (match_type, match_value) + values + (username,)
        )

    return 0


def get_transaction_splits(txn_id):
    return query(
        '''SELECT id, amount, "Признак", "Категория", "Статья", "Проект", "Контрагент_report", "Строка отчета"
           FROM flash.transaction_splits WHERE transaction_id = %s ORDER BY id''',
        (txn_id,)
    )


def set_transaction_splits(txn_id, splits, username):
    """Разбивает одну операцию на несколько частей с разными "Проект"/"Строка
    отчета" — нужно для платежей, которые в самой FinancialData распадаются
    на несколько проектов (например, оплата полевому агенту, распределённая
    по проектам его загрузки — сигнала для такого распределения нет ни в
    ИНН, ни в тексте платежа, см. RULE_CONFIDENCE_THRESHOLD выше). splits —
    список словарей с Признак/Категория/Статья/Проект/Контрагент_report/
    Строка отчета/Сумма; сумма всех частей должна совпадать с суммой самой
    операции (допуск 0.01 на округление) — иначе ValueError, чтобы случайно
    не потерять/задвоить часть платежа. Помечает исходную операцию
    classification_source='split' (из обычных отчётов Flash она после этого
    исключается, см. _effective_rows_sql) и полностью заменяет предыдущее
    разбиение этой операции, если оно было."""
    txn = query_one('SELECT amount FROM flash.transactions WHERE id = %s', (txn_id,))
    if txn is None:
        raise ValueError('Операция не найдена')
    if not splits:
        raise ValueError('Нужна хотя бы одна часть разбиения')

    total = sum(float(s.get('amount') or 0) for s in splits)
    if abs(total - float(txn['amount'])) > 0.01:
        raise ValueError(
            f'Сумма частей ({total:,.2f}) не совпадает с суммой операции ({float(txn["amount"]):,.2f})'
        )

    execute('DELETE FROM flash.transaction_splits WHERE transaction_id = %s', (txn_id,))
    rows = [(
        txn_id, s.get('amount'), s.get('Признак'), s.get('Категория'), s.get('Статья'),
        s.get('Проект'), s.get('Контрагент_report'), s.get('Строка отчета'), username,
    ) for s in splits]
    execute_values(
        '''INSERT INTO flash.transaction_splits
           (transaction_id, amount, "Признак", "Категория", "Статья", "Проект", "Контрагент_report",
            "Строка отчета", created_by)
           VALUES %s''',
        rows
    )
    execute(
        '''UPDATE flash.transactions SET "Признак"=NULL, "Категория"=NULL, "Статья"=NULL, "Проект"=NULL,
           "Контрагент_report"=NULL, "Строка отчета"=NULL, classification_source='split' WHERE id = %s''',
        (txn_id,)
    )


def clear_transaction_splits(txn_id):
    """Отменяет разбиение — операция возвращается в 'unmatched' (обычная
    точечная разметка через set_manual_classification/автоправила дальше как
    для любой другой неразмеченной операции)."""
    execute('DELETE FROM flash.transaction_splits WHERE transaction_id = %s', (txn_id,))
    execute(
        '''UPDATE flash.transactions SET classification_source='unmatched' WHERE id = %s''',
        (txn_id,)
    )


_MONTH_NAMES_RU = [
    'январь', 'февраль', 'март', 'апрель', 'май', 'июнь',
    'июль', 'август', 'сентябрь', 'октябрь', 'ноябрь', 'декабрь',
]


def export_for_load(period):
    """Строки Flash за period в точности в формате листа "загрузка"
    (monthly_etl.FACT_COLUMNS) — готовый файл для /admin/monthly_load.
    Только уже размеченные операции (classification_source != 'unmatched');
    неразмеченные не экспортируются, их нужно сначала разобрать на /flash.

    "Контрагент" берём из сырого banковского counterparty_name (не
    Контрагент_report — это уже производное отчётное имя, его пересчитает
    сам конвейер по тем же правилам, что и для остальной FinancialData).
    "Проект" копируется из последней совпавшей исторической строки
    FinancialData и не проверяется построчно — распределение по проектам
    для новых операций это допущение, а не факт; стоит перепроверить перед
    финальной загрузкой, если контрагент участвует в нескольких проектах.
    Разбитые вручную по проектам операции (см. set_transaction_splits) идут
    отдельными строками — по одной на каждую часть разбиения."""
    rows = query(
        f'''SELECT operation_date, "Признак", "Категория", "Статья", "Проект", counterparty_name,
                  wallet, amount, purpose_text
           FROM ({_effective_rows_sql()}) AS ft
           WHERE date_trunc('month', operation_date) = %s AND classification_source != 'unmatched'
           ORDER BY operation_date''',
        (period,)
    )
    wallet_types = {r['account_number']: r['wallet_type'] for r in query(
        'SELECT account_number, wallet_type FROM flash.wallet_aliases'
    )}
    accounts_by_wallet = {}
    for r in query('SELECT account_number, wallet FROM flash.wallet_aliases'):
        accounts_by_wallet[r['wallet']] = r['account_number']

    out = []
    for r in rows:
        d = r['operation_date']
        account = accounts_by_wallet.get(r['wallet'])
        out.append({
            'Дата': d, 'Год': d.year, 'месяц': _MONTH_NAMES_RU[d.month - 1],
            'Признак': r['Признак'], 'Категория': r['Категория'], 'Статья': r['Статья'],
            'Проект': r['Проект'], 'Контрагент': r['counterparty_name'],
            'Тип Кошелька': wallet_types.get(account), 'Кошелек': r['wallet'],
            'Сумма': float(r['amount']), 'Комментарии': r['purpose_text'],
        })
    return out
