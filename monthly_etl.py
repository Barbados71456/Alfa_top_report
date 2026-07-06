"""Автозагрузка отчётности: воспроизводит вручную запускаемый через KNIME-десктоп
процесс (см. /Users/borisfedulov/.claude/plans/inherited-twirling-dawn.md) как
шаги внутри приложения — сначала на отдельной схеме `etl_load`, со сверкой
против `public` перед тем, как это когда-либо тронет прод-данные.

SQL-шаги перенесены из KNIME (Alfa_для_Claude.knwf) почти дословно — тексты
лежат в etl_sql/*.sql. Ровно три места сознательно отступают от буквального
SQL узла (см. комментарии прямо в файлах 02/04/12_*.sql) — везде вместо
TRUNCATE всей таблицы используется точечный DELETE, чтобы при загрузке файла
только за один месяц не терялась история остальных месяцев. Бонусы (шаг 16)
считаются по формуле (см. ниже), а не вбиваются вручную, как в KNIME.
"""
import datetime
import json
import pathlib

import openpyxl

from db import execute, execute_returning, execute_values, query, query_one
import pl_report as pr

SQL_DIR = pathlib.Path(__file__).parent / 'etl_sql'

# Таблицы конвейера — при подстановке схемы public.<table> заменяется на
# <schema>.<table>. Справочники (dim_level_report, dim_ШР, dim_ОД_расп,
# plan_post_fin, Forecast_fin) в этот список сознательно не входят — это общие
# для факта и плана статичные данные, которые всегда читаются из public,
# конвейер их не пересобирает и не зеркалит по схемам.
PIPELINE_TABLES = [
    'FinancialData_GM', 'FinancialData', 'fact_step_0', 'fact_step_1',
    'base_report_park1', 'report_do_rasp', 'base_driver', 'rasp_FOT',
    'rasp_Ostatki', 'analiz_zaim', 'analiz_кошелек',
]

# Колонки листа "загрузка" Excel-файла = колонки public.fact_step_0.
FACT_COLUMNS = [
    'Дата', 'Год', 'месяц', 'Признак', 'Категория', 'Статья', 'Проект',
    'Контрагент', 'Тип Кошелька', 'Кошелек', 'Сумма', 'Комментарии',
]

_MONTH_NAMES = {
    'январь': 1, 'февраль': 2, 'март': 3, 'апрель': 4, 'май': 5, 'июнь': 6,
    'июль': 7, 'август': 8, 'сентябрь': 9, 'октябрь': 10, 'ноябрь': 11, 'декабрь': 12,
}

BONUS_LINES = pr.BONUS_LINES  # ['Бонус Генерального директора', 'Бонус Руководитель взыскания']
CEO_BONUS_RATE = 0.03
HEAD_COLLECTIONS_BONUS_RATE = 0.015


def _retarget(sql, schema):
    """public.xxx / public."xxx" -> <schema>.xxx / <schema>."xxx", только для
    таблиц самого конвейера (PIPELINE_TABLES). Сортировка по убыванию длины
    имени — чтобы "FinancialData" не задело "FinancialData_GM" раньше времени."""
    if schema == 'public':
        return sql
    for table in sorted(PIPELINE_TABLES, key=len, reverse=True):
        sql = sql.replace(f'public."{table}"', f'{schema}."{table}"')
        sql = sql.replace(f'public.{table}', f'{schema}.{table}')
    return sql


def _read_sql(filename):
    return (SQL_DIR / filename).read_text(encoding='utf-8')


def _prepare_sql(sql):
    """Экранирует буквальные "%" (LIKE/ILIKE-шаблоны из исходного KNIME-SQL,
    например "Проект" like '%вакуатор') под psycopg2 — иначе он пытается
    распарсить любой "%" как формат-спецификатор и падает (та же ловушка,
    что уже встречалась с ILIKE '%кошелек%' в db.query()). Единственный
    настоящий плейсхолдер %(period)s сохраняется нетронутым."""
    sentinel = '\x00PERIOD\x00'
    sql = sql.replace('%(period)s', sentinel)
    sql = sql.replace('%', '%%')
    return sql.replace(sentinel, '%(period)s')


def _month_num(raw):
    if raw is None:
        return None
    s = str(raw).strip().lower()
    if s in _MONTH_NAMES:
        return _MONTH_NAMES[s]
    try:
        n = int(float(s))
        if 1 <= n <= 12:
            return n
    except ValueError:
        pass
    return None


def parse_excel(file_obj):
    """Читает лист "загрузка" (те же 12 колонок, что в public.fact_step_0).
    Возвращает (rows, period) — период определяется по единственному
    встречающемуся сочетанию Год+месяц (файл должен быть за один месяц)."""
    wb = openpyxl.load_workbook(file_obj, data_only=True, read_only=True)
    if 'загрузка' not in wb.sheetnames:
        raise ValueError('На листе нет вкладки "загрузка"')
    ws = wb['загрузка']
    header = [str(c).strip() if c is not None else c for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    col_idx = {}
    for name in FACT_COLUMNS:
        if name not in header:
            raise ValueError(f'В файле нет колонки "{name}"')
        col_idx[name] = header.index(name)

    rows = []
    periods = set()
    for raw_row in ws.iter_rows(min_row=2, values_only=True):
        if raw_row is None or all(v is None for v in raw_row):
            continue
        rec = {name: raw_row[col_idx[name]] for name in FACT_COLUMNS}
        rows.append(rec)
        m = _month_num(rec.get('месяц'))
        god = rec.get('Год')
        if m and god:
            periods.add((int(god), m))

    if not rows:
        raise ValueError('На листе "загрузка" нет строк с данными')
    if len(periods) != 1:
        raise ValueError(f'Файл должен быть за один месяц, а найдено периодов: {sorted(periods)}')
    god, m = periods.pop()
    period = datetime.date(god, m, 1)
    return rows, period


def load_fact_step_0(schema, rows):
    """Полностью заменяет {schema}.fact_step_0 содержимым файла (сама таблица
    хранит только "сырые" данные последней загрузки, история копится дальше
    по цепочке — см. _retarget'нутый 04_financial_data_base.sql). Один
    round-trip на весь файл (execute_values), а не строка за строкой — на
    удалённой Render Postgres это разница в разы на файле в тысячи строк."""
    execute(f'TRUNCATE TABLE {schema}.fact_step_0')
    # месяц — единственная строчная колонка без пробелов и заглавных букв в
    # исходном SQL (см. etl_sql/01_fact_step1.sql), остальные квалифицируем кавычками.
    cols_sql = ', '.join('месяц' if c == 'месяц' else f'"{c}"' for c in FACT_COLUMNS)
    sql = f'INSERT INTO {schema}.fact_step_0 ({cols_sql}) VALUES %s'
    values = [tuple(r[c] for c in FACT_COLUMNS) for r in rows]
    execute_values(sql, values)
    return len(rows)


def compute_profit(schema, period):
    """Повторяет формулу прибыли из pl_report.svod1() (Итого выручка + Итого
    переменные + Итого постоянные), но напрямую по {schema}."FinancialData"
    (сумма в рублях, без allocation-фильтра — те же условия, что дефолт
    svod1(allocation='all'), это и есть база для бонусов в исходном узле 194)."""
    lines = pr.REVENUE_LINES + pr.VARIABLE_LINES + pr.FIXED_LINES
    row = query_one(
        f'''SELECT COALESCE(SUM("Сумма"), 0) AS total
           FROM {schema}."FinancialData"
           WHERE "Период" = %s AND "п_ф" = 'факт' AND "Строка отчета" = ANY(%s)''',
        (period, lines)
    )
    return row['total']


CEO_NAME = 'Мезенцев Сергей Сергеевич'
HEAD_COLLECTIONS_NAME = 'Кожевников Антон Алексеевич'

# По решению пользователя: строка бонуса — это просто расчётная величина от
# прибыли, а не классифицированная проводка (готовой строки "Бонус..." нет ни
# в public, ни в истории — узел 194 в KNIME ни разу не запускался в проде,
# у обоих людей "Строка отчета" в их реальных ЗП/премиях = NULL). Поэтому не
# клонируем несуществующий шаблон, а собираем строку сами: только то, что
# реально нужно для агрегации (Период/Распределение/п_ф/Строка отчета/Сумма),
# остальные классификационные колонки (Статья, СтатьяСвод, СтатьяУровень0-4,
# Тип Кошелька и т.п.) оставляем NULL — это сознательно держит бонус вне
# GM/OIBDA-агрегации в FinancialData_GM (она группирует по СтатьяСвод), в
# точности как в pl_report.svod1(), где БОНУСЫ — отдельная секция ПОСЛЕ
# прибыли, а не часть Выручка/Переменные/Постоянные.
def apply_bonuses(schema, period):
    """Отступление от узла 194: вместо ручной вставки суммы — считаем 3%/1.5%
    от прибыли месяца (см. compute_profit). Сумма — отрицательная (расход),
    как и остальные ФОТ/расходные строки в этой таблице."""
    profit = compute_profit(schema, period)
    ceo_bonus = -round(profit * CEO_BONUS_RATE, 2)
    head_bonus = -round(profit * HEAD_COLLECTIONS_BONUS_RATE, 2)

    execute(
        f'''DELETE FROM {schema}."FinancialData"
           WHERE "Период" = %s AND "Строка отчета" = ANY(%s)''',
        (period, BONUS_LINES)
    )

    cols = ['Распределение', 'Период', 'Признак', 'Категория', 'Проект',
            'Контрагент', 'Контрагент_report', 'п_ф', 'Строка отчета', 'Сумма']
    cols_sql = ', '.join(f'"{c}"' for c in cols)
    placeholders = ', '.join(['%s'] * len(cols))
    sql = f'INSERT INTO {schema}."FinancialData" ({cols_sql}) VALUES ({placeholders})'

    for line, contragent, amount in (
        (BONUS_LINES[0], CEO_NAME, ceo_bonus),
        (BONUS_LINES[1], HEAD_COLLECTIONS_NAME, head_bonus),
    ):
        row = {
            'Распределение': 'до распределения', 'Период': period, 'Признак': 'финансы',
            'Категория': 'Бонус', 'Проект': 'Общее', 'Контрагент': contragent,
            'Контрагент_report': 'Дирекция', 'п_ф': 'факт', 'Строка отчета': line, 'Сумма': amount,
        }
        execute(sql, tuple(row[c] for c in cols))
    return {'profit': profit, 'ceo_bonus': ceo_bonus, 'head_bonus': head_bonus}


# Шаги конвейера в порядке восстановленного графа связей workflow.knime.
# (id, label, sql_filename или None для питон-шагов)
STEPS = [
    ('fact_step_1', 'Fact step 1 (перенос сырых данных)', '01_fact_step1.sql'),
    ('base_report_park1', 'Base report park1', '02_base_report_park1.sql'),
    ('report_do_rasp', 'Report до распределения (классификация статей)', '03_report_do_rasp.sql'),
    ('financial_data_base', 'FinancialData: факт "до распределения"', '04_financial_data_base.sql'),
    ('classify_1', 'Доклассификация 1/4', '05_classify_1.sql'),
    ('classify_2', 'Доклассификация 2/4', '06_classify_2.sql'),
    ('classify_3', 'Доклассификация 3/4', '07_classify_3.sql'),
    ('classify_4', 'Доклассификация 4/4', '08_classify_4.sql'),
    ('base_driver', 'Драйверы распределения затрат', '09_base_driver.sql'),
    ('rasp_fot', 'Распределение ФОТ по проектам', '10_rasp_fot.sql'),
    ('rasp_ostatki', 'Распределение прочих расходов по проектам', '11_rasp_ostatki.sql'),
    ('financial_data_allocated', 'FinancialData: строки "распределение"', '12_financial_data_allocated.sql'),
    ('analiz_zaim', 'Анализ займов', '13_analiz_zaim.sql'),
    ('analiz_koshelek', 'Анализ кошельков', '14_analiz_koshelek.sql'),
    ('financial_data_gm', 'FinancialData_GM (GM/OIBDA/прибыль)', '15_financial_data_gm.sql'),
]


def run_pipeline(schema, file_obj, started_by):
    """Прогоняет весь конвейер на указанной схеме. Возвращает
    (period, steps) — steps это список {step, label, status, detail}."""
    steps_result = []

    def _step(step_id, label, fn):
        try:
            detail = fn()
            steps_result.append({'step': step_id, 'label': label, 'status': 'ok', 'detail': detail})
        except Exception as e:
            steps_result.append({'step': step_id, 'label': label, 'status': 'error', 'detail': str(e)})
            raise

    period = None
    run_id = None
    try:
        rows, period = parse_excel(file_obj)
        run_id = _log_run_start(period, started_by)

        _step('excel', 'Чтение Excel-файла (лист "загрузка")', lambda: f'{len(rows)} строк, период {period}')
        _step('fact_step_0', 'Загрузка в fact_step_0', lambda: f'{load_fact_step_0(schema, rows)} строк')

        for step_id, label, filename in STEPS:
            sql = _prepare_sql(_retarget(_read_sql(filename), schema))
            _step(step_id, label, lambda sql=sql, period=period: _run_sql(sql, period))

        _step('bonuses', 'Бонусы (3% ГД / 1.5% рук. взыскания от прибыли месяца)',
              lambda: _format_bonus_detail(apply_bonuses(schema, period)))
    finally:
        if run_id is not None:
            _log_run_finish(run_id, steps_result)

    return period, steps_result


def _format_bonus_detail(b):
    fmt = lambda v: f'{v:,.2f}'.replace(',', ' ')
    return (f'Прибыль {fmt(b["profit"])} ₽; бонус ГД {fmt(b["ceo_bonus"])} ₽; '
            f'бонус рук. взыскания {fmt(b["head_bonus"])} ₽')


def _run_sql(sql, period):
    execute(sql, {'period': period})
    return 'ok'


def _log_run_start(period, started_by):
    row = execute_returning(
        '''INSERT INTO etl_load.run_log (period, started_by, status) VALUES (%s, %s, 'running') RETURNING id''',
        (period, started_by)
    )
    return row['id']


def _log_run_finish(run_id, steps_result):
    status = 'error' if any(s['status'] == 'error' for s in steps_result) else 'ok'
    execute(
        'UPDATE etl_load.run_log SET steps = %s, status = %s WHERE id = %s',
        (json.dumps(steps_result, default=str), status, run_id)
    )


def compare_with_public(schema, period):
    """Сравнивает суммы по "Строка отчета" между {schema}."FinancialData" и
    public."FinancialData" за один и тот же период. Возвращает список
    расхождений {line, schema_val, public_val, diff} (пустой список = 1-в-1).

    Сравнение сознательно ограничено "Распределение"='до распределения' —
    это классифицированный факт из Excel (единственное, что можно сравнить
    1-в-1 построчно, т.к. оно не зависит от истории других месяцев). Секцию
    "распределение" (косвенные затраты) сравнивать некорректно: в public она
    сейчас вообще пустая (0 строк на момент внедрения — видимо, после
    какой-то из прошлых полных перезаливок эти строки потерялись и не были
    пересчитаны), а строится она из ВСЕЙ истории "до распределения", которой
    в {schema} может не хватать, если туда загружен только один месяц."""
    schema_rows = {
        r['line']: r['val'] for r in query(
            f'''SELECT "Строка отчета" AS line, SUM("Сумма") AS val
               FROM {schema}."FinancialData"
               WHERE "Период" = %s AND "Распределение" = 'до распределения' GROUP BY 1''',
            (period,)
        )
    }
    public_rows = {
        r['line']: r['val'] for r in query(
            '''SELECT "Строка отчета" AS line, SUM("Сумма") AS val
               FROM public."FinancialData"
               WHERE "Период" = %s AND "Распределение" = 'до распределения' GROUP BY 1''',
            (period,)
        )
    }
    diffs = []
    for line in sorted(set(schema_rows) | set(public_rows)):
        sv = schema_rows.get(line, 0) or 0
        pv = public_rows.get(line, 0) or 0
        if abs(sv - pv) > 0.01:
            diffs.append({'line': line, 'schema_val': sv, 'public_val': pv, 'diff': sv - pv})
    return diffs


def get_run_log(limit=20):
    return query('SELECT * FROM etl_load.run_log ORDER BY started_at DESC LIMIT %s', (limit,))
