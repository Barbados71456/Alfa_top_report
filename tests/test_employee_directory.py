from io import BytesIO
import unittest
from unittest.mock import MagicMock, patch

from openpyxl import load_workbook

from employee_directory import (
    EmployeeWorkbookError,
    apply_employee_updates,
    build_employee_summary,
    build_employee_workbook,
    parse_employee_workbook,
)


class EmployeeWorkbookTests(unittest.TestCase):
    def setUp(self):
        self.rows = [
            {
                'contragent': 'Иванов Иван Иванович',
                'department': 'Продажи',
                'position': 'Менеджер',
                'status': 'Работает',
            },
            {
                'contragent': 'Петрова Анна Сергеевна',
                'department': None,
                'position': None,
                'status': 'Уволен',
            },
        ]

    def test_export_has_expected_structure_and_round_trips(self):
        output = build_employee_workbook(self.rows, ['Продажи', 'Маркетинг'])
        payload = output.getvalue()
        workbook = load_workbook(BytesIO(payload))
        self.assertEqual(workbook.sheetnames, ['Сотрудники', 'Справочники', 'Инструкция'])
        self.assertEqual(workbook['Справочники'].sheet_state, 'hidden')
        self.assertEqual(
            [workbook['Сотрудники'].cell(1, column).value for column in range(1, 5)],
            ['Контрагент', 'Подразделение', 'Должность', 'Статус'],
        )
        self.assertEqual(workbook['Сотрудники'].freeze_panes, 'A2')
        validations = workbook['Сотрудники'].data_validations.dataValidation
        self.assertGreaterEqual(len(validations), 2)
        self.assertEqual({item.formula1 for item in validations}, {
            'EmployeeDepartments', 'EmployeeStatuses',
        })
        self.assertIn('EmployeeDepartments', workbook.defined_names)
        self.assertIn('EmployeeStatuses', workbook.defined_names)
        workbook.close()

        parsed = parse_employee_workbook(payload)
        self.assertEqual(parsed, self.rows)

    def test_summary_groups_by_department_and_status_with_totals(self):
        rows = self.rows + [
            {
                'contragent': 'Сидоров Пётр Олегович',
                'department': 'Продажи',
                'position': 'Руководитель',
                'status': None,
            },
            {
                'contragent': 'Ким Анна Олеговна',
                'department': 'Архив',
                'position': None,
                'status': 'Отпуск',
            },
        ]

        summary = build_employee_summary(rows, ['Продажи', 'Маркетинг'])

        self.assertEqual(summary['statuses'], ['Работает', 'Уволен', 'Отпуск'])
        self.assertEqual(
            [row['department'] for row in summary['rows']],
            ['Продажи', 'Архив', 'Без подразделения'],
        )
        self.assertEqual(summary['rows'][0]['counts'], {
            'Работает': 2, 'Уволен': 0, 'Отпуск': 0,
        })
        self.assertEqual(summary['totals'], {
            'Работает': 2, 'Уволен': 1, 'Отпуск': 1,
        })
        self.assertEqual(summary['grand_total'], 4)

    def test_summary_handles_empty_list(self):
        summary = build_employee_summary([], ['Продажи'])

        self.assertEqual(summary['rows'], [])
        self.assertEqual(summary['totals'], {'Работает': 0, 'Уволен': 0})
        self.assertEqual(summary['grand_total'], 0)

    def test_import_rejects_duplicate_employee(self):
        rows = [self.rows[0], dict(self.rows[0])]
        with self.assertRaisesRegex(EmployeeWorkbookError, 'повторно'):
            parse_employee_workbook(build_employee_workbook(rows).getvalue())

    def test_import_rejects_formula(self):
        workbook = load_workbook(BytesIO(build_employee_workbook(self.rows).getvalue()))
        workbook['Сотрудники']['C2'] = '=1+1'
        output = BytesIO()
        workbook.save(output)
        workbook.close()
        with self.assertRaisesRegex(EmployeeWorkbookError, 'формулы запрещены'):
            parse_employee_workbook(output.getvalue())

    @patch('employee_directory.transaction')
    def test_apply_updates_only_changed_rows(self, transaction_mock):
        connection = MagicMock()
        cursor = MagicMock()
        cursor.__enter__.return_value = cursor
        cursor.__exit__.return_value = False
        cursor.fetchall.return_value = [
            {
                'contragent': 'Иванов Иван Иванович',
                'department': 'Продажи',
                'position': 'Стажёр',
                'status': 'Работает',
            },
            {
                'contragent': 'Петрова Анна Сергеевна',
                'department': None,
                'position': None,
                'status': 'Уволен',
            },
        ]
        connection.cursor.return_value = cursor
        transaction_mock.return_value.__enter__.return_value = connection
        transaction_mock.return_value.__exit__.return_value = False

        result = apply_employee_updates(self.rows)

        self.assertEqual(result, {'processed': 2, 'updated': 1, 'unchanged': 1})
        cursor.executemany.assert_called_once()
        parameters = cursor.executemany.call_args.args[1]
        self.assertEqual(parameters, [('Продажи', 'Менеджер', 'Работает', 'Иванов Иван Иванович')])

    @patch('employee_directory.transaction')
    def test_apply_rejects_missing_employee_before_update(self, transaction_mock):
        connection = MagicMock()
        cursor = MagicMock()
        cursor.__enter__.return_value = cursor
        cursor.__exit__.return_value = False
        cursor.fetchall.return_value = []
        connection.cursor.return_value = cursor
        transaction_mock.return_value.__enter__.return_value = connection
        transaction_mock.return_value.__exit__.return_value = False

        with self.assertRaisesRegex(EmployeeWorkbookError, 'Не найдены сотрудники'):
            apply_employee_updates([self.rows[0]])
        cursor.executemany.assert_not_called()


if __name__ == '__main__':
    unittest.main()
