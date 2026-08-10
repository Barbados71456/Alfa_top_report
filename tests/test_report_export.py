from datetime import date
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

import export
import fot_report as fr


REPO_ROOT = Path(__file__).resolve().parents[1]


def test_workbook_has_consistent_excel_formatting_and_safe_values():
    payload = export.build_workbook([
        (
            'Отчёт/Продажи',
            ['Статья', 'Дата', 'Сумма, руб', 'Маржа, %', 'Комментарий'],
            [
                ['Выручка', date(2026, 8, 1), 123456.75, 12.5, '=HYPERLINK("https://example.com")'],
                ['Расходы', date(2026, 8, 2), -456.5, -2.25, 'Проверено'],
            ],
        ),
        ('Отчёт:Продажи', ['Показатель'], [['Итого']]),
    ]).getvalue()

    workbook = load_workbook(BytesIO(payload), data_only=False)
    assert workbook.sheetnames == ['Отчёт Продажи', 'Отчёт Продажи (2)']
    sheet = workbook['Отчёт Продажи']
    assert sheet.freeze_panes == 'B2'
    assert sheet.auto_filter.ref == 'A1:E3'
    assert sheet.sheet_view.showGridLines is False
    assert sheet['A1'].font.bold is True
    assert sheet['A1'].font.color.rgb == '00FFFFFF'
    assert sheet['A1'].fill.fgColor.rgb == '001F4E78'
    assert sheet['B2'].number_format == export.DATE_FORMAT
    assert sheet['C2'].number_format == export.NUMBER_FORMAT
    assert sheet['D2'].number_format == export.PERCENT_FORMAT
    assert sheet['E2'].data_type == 's'
    assert sheet['E2'].value.startswith("'=")
    assert sheet.column_dimensions['A'].width >= len('Статья')
    workbook.close()


def test_fot3_export_contains_employee_parameters_and_totals():
    variable = [float(value) for value in range(1, 13)]
    fixed = [float(value * 10) for value in range(1, 13)]
    total = [variable[index] + fixed[index] for index in range(12)]
    data = {
        'employee': 'Иванов И.И.',
        'dept': 'Commercial',
        'months': fr.MONTHS_RU,
        'variable': variable,
        'fixed': fixed,
        'total': total,
        'year_total': sum(total),
    }

    payload = export.build_workbook(fr.export_fot3(data, year=2026, pf='факт')).getvalue()
    workbook = load_workbook(BytesIO(payload), data_only=True)
    assert workbook.sheetnames == ['ФОТ v3', 'Параметры']
    report = workbook['ФОТ v3']
    assert report.max_column == 14
    assert report['A2'].value == 'ФОТ переменный'
    assert report['N4'].value == sum(total)
    parameters = workbook['Параметры']
    assert parameters['B2'].value == 'Иванов И.И.'
    assert parameters['B3'].value == 'Commercial'
    assert parameters['B4'].value == 2026
    assert parameters['B5'].value == 'факт'
    workbook.close()


def test_every_report_template_exposes_excel_export():
    report_templates = [
        'overview.html',
        'wallets_summary.html',
        'wallets_detail.html',
        'wallets_ledger.html',
        'flash.html',
        'svod1.html',
        'svod2.html',
        'dashboard1.html',
        'dashboard2.html',
        'unitpl.html',
        'fot1.html',
        'fot2.html',
        'fot3.html',
        'loans.html',
        'counterparty.html',
        'investment_summary.html',
        'investment_detail.html',
        'cbr.html',
    ]

    for template_name in report_templates:
        template = (REPO_ROOT / 'templates' / template_name).read_text(encoding='utf-8')
        assert "export_report" in template, f'Нет Excel-экспорта в {template_name}'
